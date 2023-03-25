from docxtpl import DocxTemplate
from openpyxl import load_workbook
from pathlib import Path

current_dir = Path.cwd()
doc = DocxTemplate('template.docx')
workbook = load_workbook('info_table.xlsx')
worksheet = workbook.active
# print(worksheet.max_row)
for number in range(3,worksheet.max_row+1):  #worksheet.max_row+1
    value_list = []
    for row in worksheet.iter_rows(min_row=number,max_row=number+1):
        for cell in row:
            # seria = cell[0].value
            value_list.append(cell.value)
            # print(cell.value)
    # print(value_list)
    # seria = value_list[0]
    serial_number = value_list[1]
    company_name = value_list[2]
    owner = value_list[3]
    credit_code = value_list[4]
    company_address = value_list[5]
    company_telephone = value_list[6]
    employee_name = value_list[7]
    gender = value_list[8]
    ID_number = value_list[9]
    employee_address = value_list[10]
    employee_telephone = value_list[11]
    emergency_telephone = value_list[12]
    start_year = value_list[13]
    start_month = value_list[14]
    start_day = value_list[15]
    end_year = value_list[16]
    end_month = value_list[17]
    end_day = value_list[18]
    probation = value_list[19]
    job_name = value_list[20]
    salary = value_list[21]
    store_name = value_list[22]
    board_date = value_list[23]


    data = {}
    data['serial_number'] = serial_number
    data['company_name'] = company_name
    data['owner'] = owner
    data['credit_code'] = credit_code
    data['company_address'] = company_address
    data['company_telephone'] = company_telephone
    data['employee_name'] = employee_name
    data['gender'] = gender
    data['ID_number'] = ID_number
    data['employee_address'] = employee_address
    data['employee_telephone'] = employee_telephone
    data['emergency_telephone'] = emergency_telephone
    data['start_year'] = start_year
    data['start_month'] = start_month
    data['start_day'] = start_day
    data['end_year'] = end_year
    data['end_month'] = end_month
    data['end_day'] = end_day
    data['probation'] = probation
    data['job_name'] = job_name
    data['salar'] = salary
    data['board_date'] = board_date

    path = current_dir.joinpath(store_name)
    if not path.exists():
        path.mkdir()


    doc.render(data)
    doc.save(f'./{store_name}/{store_name}-{employee_name}.docx')